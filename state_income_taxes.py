# -*- coding: utf-8 -*-
"""
Created on Mon May 24 15:03:46 2021

@author: One
"""
from openpyxl import load_workbook

import pandas as pd
import os
#takes in tax_rates (created from create_dataframe()), the national_sample_census, and year
def calculate_average(tax_rates, df, year):
    df['Average Tax Rate'] = df.apply(lambda row: average_tax(tax_rates, row['statename'], row['pc_income_census2010'], year, 'rate'), axis = 1)
    df['Average Tax Amount'] = df.apply(lambda row: average_tax(tax_rates, row['statename'], row['pc_income_census2010'], year,'amount'), axis = 1)
    return df
def average_tax(tax_rates,state, gross_income, year, return_type):
    if year < 2015:
        year = 2015
    states = ['AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY', 'DC']
    states_full = ['Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut', 'Delaware', 'Florida', 'Georgia',  'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia', 'Wisconsin', 'Wyoming','DC']
    states_dict = dict(zip(states_full,states))
    if len(state) > 2:
        state = states_dict[state]
    #Joint #One dependent
    tax_rates = tax_rates[tax_rates['State'] == state ]
    tax_rates = tax_rates[tax_rates['Year'] == str(year) ]
    #tax_rates = tax_rates[tax_rates['Year'] == str(year) ]
    tax_rates = tax_rates[tax_rates['File Type'] == 'Couple']
    rates = list(tax_rates['Rate'])
    brackets = list(tax_rates['Bracket'])
    
    
    s_row = tax_rates.head(1)
    
    taxable_income_reduction = 0
    credit = 0
    #print(s_row['Dependent Exemption Type'].values)
    if s_row['Dependent Exemption Type'].values[0] == False:
        if s_row['Dependent Exemption'].values[0] not in ['n.a.','n.a','(l)']:
            taxable_income_reduction += s_row['Dependent Exemption'].values[0]
    else:
        credit += s_row['Dependent Exemption'].values[0]
    
    if s_row['Exemption Type'].values[0] == False:
        if s_row['Exemption'].values[0] not in ['n.a.','n.a','(l)']:
            taxable_income_reduction += s_row['Exemption'].values[0]
    else:
        credit += s_row['Exemption'].values[0]
        
    if s_row['Deduction Type'].values[0] == False:
        if s_row['Deduction'].values[0] not in ['n.a.','n.a','(l)']:
            taxable_income_reduction += s_row['Deduction'].values[0]
    else:
        credit += s_row['Deduction'].values[0]
    
    taxed_amount = 0
    if 'none' in brackets or 'none' in rates:
        return 0
    income = gross_income - taxable_income_reduction
    if income < 0:
        income = 0
    #print(income,rates)
    for i,bracket in enumerate(brackets):
        if len(brackets) > i + 1:
            if income > brackets[i+1]:
                taxed_amount += (brackets[i+1] - brackets[i])*rates[i]
                print((brackets[i+1] - brackets[i])*rates[i])
                
        #need to figure this out better
            else:
                if income > brackets[i]:
                    taxed_amount += (income - brackets[i])*rates[i]
                    print((income - brackets[i])*rates[i])
                else:
                    break
        else:
            if income > brackets[i]:
                taxed_amount += (income - brackets[i])*rates[i]
                print((income - brackets[i])*rates[i])
            else:
                break
        # if income > bracket:
        #     if len(bracket) > i + 1:
        #         taxed_amount += bracket[i+1]*rates[i]
        #         previous_bracket = bracket

        # if income <= bracket:
        #     taxed_amount += (income - previous_bracket)*rates[i]
        #     break
    #print(taxed_amount, credit)
    taxed_amount = taxed_amount - credit
    average_tax_rate = taxed_amount/gross_income
    if return_type == "amount":
        return taxed_amount
    elif return_type == "rate":
        return average_tax_rate
    #return (taxed_amount, average_tax_rate)
    
def create_dataframe():
    #Change this to name of the excel sheet
    excel_name = "\State-Individual-Income-Tax-Rates-and-Brackets-for-2021.xlsx"
    # workbook = pd.read_excel(os.getcwd() + excel_name,sheet_name = '2021')
    
    #alter this line to be where the excel file is located
    workbook = load_workbook(os.getcwd() + excel_name)
    sheet = workbook.active
    states = ['AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY', 'DC']
      
    data = []

    
    
    for x in workbook.sheetnames:
        count = 0
        sheet = workbook[x]
        print(x)
        curr_state = None
        state_counter = 0
        
        
        for row_index, row in enumerate(sheet.iter_rows()):
            deduction = {}
            exemption = {}
            curr_state = None
            
            credit = {}
            if row[0].value != None:
                #this is to not allow it to read on forever
                if len(row[0].value) > 24:
                    break
                if not row[0].value.islower() and row[0].value!= "State":
                    count += 1
                    curr_state = row[0].value
                    #Override with better format
                    curr_state = states[state_counter]
                    #print(row[0].value, states[state_counter])
                    state_counter += 1
                    if x != '2018' and x != '2017':
                        
                        if 'credit' in str(sheet['H' + str(row_index + 1)].value):
                            credit['Deduction Single'] = True
                            for word in str(sheet['H' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    deduction['single'] = int(word)
                        else:
                            credit['Deduction Single'] = False
                            deduction['single'] = sheet['H' + str(row_index + 1)].value
                            
                            
                            
                        if 'credit' in str(sheet['I' + str(row_index + 1)].value):
                            credit['Deduction Couple'] = True
                            for word in str(sheet['I' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    deduction['couple'] = int(word)
                        else:
                            credit['Deduction Couple'] = False
                            deduction['couple'] = sheet['I' + str(row_index + 1)].value
                            
                            
                        if 'credit' in str(sheet['J' + str(row_index + 1)].value):
                            credit['Exemption Single'] = True
                            for word in str(sheet['J' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['single'] = int(word)
                        else:
                            credit['Exemption Single'] = False
                            exemption['single'] = sheet['J' + str(row_index + 1)].value
                            
                        if 'credit' in str(sheet['K' + str(row_index + 1)].value):
                            credit['Exemption Couple'] = True
                            for word in str(sheet['K' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['couple'] = int(word)
                        else:
                            credit['Exemption Couple'] = False
                            exemption['couple'] = sheet['K' + str(row_index + 1)].value
                        
                        if 'credit' in str(sheet['L' + str(row_index + 1)].value):
                            credit['Exemption Dependent'] = True
                            for word in str(sheet['L' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['dependent'] = int(word)
                        else:
                            credit['Exemption Dependent'] = False
                            exemption['dependent'] = sheet['L' + str(row_index + 1)].value
                        i = 1
                        while sheet['B' + str(row_index+i)].value != None or sheet['E' + str(row_index+i)].value != None  :
                            #single
                            if sheet['B' + str(row_index+i)].value != None:
                                row = {}
                                row['State'] = curr_state
                                row['Year'] = str(x)
                                row['File Type'] = "Single"
                                row["Rate"] = sheet['B' + str(row_index+i)].value
                                row["Bracket"] = sheet['D' + str(row_index+i)].value
                                row['Deduction'] = deduction['single']
                                row['Deduction Type'] = credit['Deduction Single']
                                row['Exemption'] = exemption['single']
                                row['Exemption Type'] = credit['Exemption Single']
                                row['Dependent Exemption'] = exemption['dependent']
                                row['Dependent Exemption Type'] = credit['Exemption Dependent']
                                data.append(row)
                            #couple
                            if sheet['E' + str(row_index+i)].value != None:
                                row = {}
                                row['State'] = curr_state
                                row['Year'] = str(x)
                                row['File Type'] = "Couple"
                                row["Rate"] = sheet['E' + str(row_index+i)].value
                                row["Bracket"] = sheet['G' + str(row_index+i)].value
                                row['Deduction'] = deduction['couple']
                                row['Deduction Type'] = credit['Deduction Couple']
                                row['Exemption'] = exemption['couple']
                                row['Exemption Type'] = credit['Exemption Couple']
                                row['Dependent Exemption'] = exemption['dependent']
                                row['Dependent Exemption Type'] = credit['Exemption Dependent']
                                data.append(row)
                            i += 1
                        
                        
                    else:
                        if 'credit' in str(sheet['J' + str(row_index + 1)].value):
                            credit['Deduction Single'] = True
                            for word in str(sheet['J' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    deduction['single'] = int(word)
                        else:
                            credit['Deduction Single'] = False
                            deduction['single'] = sheet['J' + str(row_index + 1)].value
                            
                            
                            
                        if 'credit' in str(sheet['K' + str(row_index + 1)].value):
                            credit['Deduction Couple'] = True
                            for word in str(sheet['K' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    deduction['couple'] = int(word)
                        else:
                            credit['Deduction Couple'] = False
                            deduction['couple'] = sheet['K' + str(row_index + 1)].value
                            
                            
                        if 'credit' in str(sheet['M' + str(row_index + 1)].value):
                            credit['Exemption Single'] = True
                            for word in str(sheet['M' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['single'] = int(word)
                        else:
                            credit['Exemption Single'] = False
                            exemption['single'] = sheet['M' + str(row_index + 1)].value
                            
                        if 'credit' in str(sheet['N' + str(row_index + 1)].value):
                            credit['Exemption Couple'] = True
                            for word in str(sheet['N' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['couple'] = int(word)
                        else:
                            credit['Exemption Couple'] = False
                            exemption['couple'] = sheet['N' + str(row_index + 1)].value
                        
                        if 'credit' in str(sheet['O' + str(row_index + 1)].value):
                            credit['Exemption Dependent'] = True
                            for word in str(sheet['O' + str(row_index + 1)].value).replace('$','').replace(',','').split():
                                if word.isdigit():
                                    exemption['dependent'] = int(word)
                        else:
                            credit['Exemption Dependent'] = False
                            exemption['dependent'] = sheet['O' + str(row_index + 1)].value
                        
                        i = 1
                        while sheet['B' + str(row_index+i)].value != None or sheet['E' + str(row_index+i)].value != None  :
                            #single
                            if sheet['B' + str(row_index+i)].value != None:
                                row = {}
                                row['State'] = curr_state
                                row['Year'] = str(x)
                                row['File Type'] = "Single"
                                row["Rate"] = sheet['B' + str(row_index+i)].value
                                row["Bracket"] = sheet['D' + str(row_index+i)].value
                                row['Deduction'] = deduction['single']
                                row['Deduction Type'] = credit['Deduction Single']
                                row['Exemption'] = exemption['single']
                                row['Exemption Type'] = credit['Exemption Single']
                                row['Dependent Exemption'] = exemption['dependent']
                                row['Dependent Exemption Type'] = credit['Exemption Dependent']
                                data.append(row)
                            #couple
                            if sheet['F' + str(row_index+i)].value != None:
                                row = {}
                                row['State'] = curr_state
                                row['Year'] = str(x)
                                row['File Type'] = "Couple"
                                row["Rate"] = sheet['F' + str(row_index+i)].value
                                row["Bracket"] = sheet['H' + str(row_index+i)].value
                                row['Deduction'] = deduction['couple']
                                row['Deduction Type'] = credit['Deduction Couple']
                                row['Exemption'] = exemption['couple']
                                row['Exemption Type'] = credit['Exemption Couple']
                                row['Dependent Exemption'] = exemption['dependent']
                                row['Dependent Exemption Type'] = credit['Exemption Dependent']
                                data.append(row)
                            i += 1
                    
                    
        print(count)
    
    df = pd.DataFrame(data)
    print(df)
    df.to_pickle(os.getcwd() + "\state_income_tax.pkl")
if __name__ == "__main__":
    #This currently assumes that the excel sheet for the rates and bracket is in the current directory
    #and creates state_income_tax.pkl in that directory as well
    create_dataframe()
    tr = pd.read_pickle('state_income_tax.pkl')
    #It also assumes that this pkl file below is in the same directory as well
    df_income = pd.read_pickle('national_sample_census2010_full.pkl')
    print(df_income.columns)
    
    df = calculate_average(tr,df_income, 2010)
    print(df[['statename','pc_income_census2010','Average Tax Amount','Average Tax Rate']].sort_values(by = 'Average Tax Amount'))
    #Name and directory of the created pickle
    df.to_pickle('2010_estimated_tax.pkl')
    print(df)
